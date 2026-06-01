"""
Parse the Atlanta Billing XLSX dropped into incoming-freightiq/ and produce
the ATL_BILLING constant block to paste into src/App.jsx.

Usage:
  python scripts/parse_atl_billing.py

Counts EVERY load in the sheet. The "Assigned" column has values like
'ATL', 'ASSIGNED TO CORP', 'ASSIGNED TO CEE' that reflect QBO booking
routing, but per Ben every load in the sheet is ATL revenue regardless
of where it's invoiced from. Don't filter by Assigned.

First-name → PAYROLL name mapping (extend if new ATL drivers appear):
  Anthoni → Davis Anthoni D
  Sam     → Denman Samuel E
  Michael → Wainwright Michael W
  CJ      → Johnson Christopher
  Manar   → Alshamaa Manar
  Robert  → Tucker Robert
"""
import os, sys, glob, re
import openpyxl
from collections import defaultdict

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INCOMING = os.path.join(BASE, "incoming-freightiq")

NAME_MAP = {
    "Anthoni": "Davis Anthoni D",
    "ANTHONI": "Davis Anthoni D",  # casing variant
    "Anthony": "Davis Anthoni D",  # spelling variant
    "Sam":     "Denman Samuel E",
    "Michael": "Wainwright Michael W",
    "CJ":      "Johnson Christopher",
    "Manar":   "Alshamaa Manar",
    "Robert":  "Tucker Robert",
}


def find_billing_file():
    # Accept either old name (2026-Atlanta Billing.xlsx), short name (ATL.xlsx),
    # or new naming ("ATLANTA 2026 (N).xlsx"). Pick the most recent match.
    patterns = ["*Atlanta Billing*.xlsx", "ATL.xlsx", "*ATL*.xlsx", "ATLANTA*.xlsx"]
    matches = []
    for pat in patterns:
        matches += glob.glob(os.path.join(INCOMING, pat))
    matches = list(set(matches))
    if not matches:
        print(f"No ATL billing XLSX in {INCOMING}/ (looked for 'Atlanta Billing', 'ATL.xlsx', or 'ATLANTA*.xlsx')")
        sys.exit(1)
    return max(matches, key=os.path.getmtime)


def pick_sheet(wb):
    """Prefer 'ALL LOADS THRU' consolidated sheet (new format); fall back to 'as of <date>' (old format); else sheet 0."""
    for sn in wb.sheetnames:
        if sn.upper().startswith("ALL LOADS"):
            return wb[sn]
    for sn in wb.sheetnames:
        if sn.lower().startswith("as of"):
            return wb[sn]
    return wb[wb.sheetnames[0]]


def build_col_index(header_row):
    """Map normalized header → column index. Handles trailing spaces, casing, and the
    'Load #' vs 'Load $' variants. Returns dict with keys: driver, invoice, carrier, assigned."""
    idx = {}
    for i, h in enumerate(header_row):
        if not h:
            continue
        s = str(h).strip().lower().rstrip("#$").strip()
        if s == "driver":                   idx["driver"]   = i
        elif s in ("invoice amount", "invoice"): idx["invoice"] = i
        elif s in ("carrier amount",):      idx["carrier"]  = i
        elif s == "carrier":
            # The "Carrier" column (just the name) sits BEFORE "Carrier Amount".
            # We want the amount, not the name — don't overwrite if already set.
            idx.setdefault("_carrier_name", i)
        elif s in ("assigned", "office"):   idx["assigned"] = i
    # If we still have no "carrier" amount, we couldn't find it — caller will error.
    return idx


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = pick_sheet(wb)

    # Extract "as of <date>" from sheet name or fall back to file mtime
    as_of_match = re.search(r"(?:as of|THRU)\s*(\d+)[-./](\d+)(?:[-./](\d+))?", sh.title, re.IGNORECASE)
    if as_of_match:
        m, d, y = as_of_match.groups()
        months = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        year = f"20{y}" if y else "2026"
        as_of = f"{months[int(m)]} {int(d)}, {year}"
    else:
        import datetime
        mtime = datetime.datetime.fromtimestamp(os.path.getmtime(path))
        as_of = mtime.strftime("%b %-d, %Y") if os.name != "nt" else mtime.strftime("%b %#d, %Y")

    rows = list(sh.iter_rows(values_only=True))
    cols = build_col_index(rows[0])
    required = {"driver", "invoice", "carrier"}
    missing = required - cols.keys()
    if missing:
        print(f"ERROR: required columns missing from sheet '{sh.title}': {missing}")
        print(f"Headers found: {rows[0]}")
        sys.exit(1)

    atl_rows = []
    assigned_counts = defaultdict(int)
    for r in rows[1:]:
        if not r:
            continue
        driver_val = r[cols["driver"]] if cols["driver"] < len(r) else None
        if not driver_val:
            continue
        assigned_idx = cols.get("assigned")
        assigned = (str(r[assigned_idx]).strip().upper() if assigned_idx is not None and assigned_idx < len(r) and r[assigned_idx] else "")
        assigned_counts[assigned] += 1
        # Repackage into a fixed-position tuple so downstream code stays simple:
        # (driver, _, _, _, _, invoice_amount, _, carrier_amount, assigned)
        invoice = r[cols["invoice"]] if cols["invoice"] < len(r) else None
        carrier = r[cols["carrier"]] if cols["carrier"] < len(r) else None
        atl_rows.append((driver_val, None, None, None, None, invoice, None, carrier, assigned))

    # Aggregate by FULL NAME (after NAME_MAP) so casing/spelling variants roll up
    by_full = defaultdict(lambda: {"count": 0, "revenue": 0.0, "carrier": 0.0, "shorts": set()})
    total_rev = 0.0
    total_car = 0.0
    for r in atl_rows:
        driver_short = (str(r[0]).strip() if r[0] else "")
        full_name = NAME_MAP.get(driver_short, f"<UNMAPPED: {driver_short}>")
        invoice = float(r[5]) if isinstance(r[5], (int, float)) else 0
        carrier = float(r[7]) if isinstance(r[7], (int, float)) else 0
        by_full[full_name]["count"] += 1
        by_full[full_name]["revenue"] += invoice
        by_full[full_name]["carrier"] += carrier
        by_full[full_name]["shorts"].add(driver_short)
        total_rev += invoice
        total_car += carrier
    # Map back to driver_short-keyed dict for emit (use first short variant as the short label)
    by_driver = {}
    for full_name, t in by_full.items():
        # Use the canonical short (first entry in NAME_MAP that maps to this full_name) if known
        canonical_short = next((k for k, v in NAME_MAP.items() if v == full_name and k[0].isupper() and k[1:].islower()), None)
        if canonical_short is None:
            canonical_short = sorted(t["shorts"])[0] if t["shorts"] else ""
        by_driver[canonical_short] = {"count": t["count"], "revenue": t["revenue"], "carrier": t["carrier"]}

    return {
        "as_of": as_of,
        "loads": len(atl_rows),
        "revenue": round(total_rev, 2),
        "carrier_pay": round(total_car, 2),
        "gross_profit": round(total_rev - total_car, 2),
        "gross_margin": round((total_rev - total_car) / total_rev * 100, 1) if total_rev else 0,
        "by_driver": by_driver,
        "assigned_counts": dict(assigned_counts),
    }


def emit(data):
    print("-" * 60)
    print(f"  ATL BILLING — as of {data['as_of']}")
    print("-" * 60)
    print(f"  ATL loads:    {data['loads']}")
    print(f"  Revenue:      ${data['revenue']:>12,.2f}")
    print(f"  Carrier pay:  ${data['carrier_pay']:>12,.2f}")
    print(f"  Gross profit: ${data['gross_profit']:>12,.2f}")
    print(f"  Gross margin: {data['gross_margin']}%")
    print()
    print("  Assigned breakdown (all loads count as ATL revenue per Ben):")
    for k, n in sorted(data["assigned_counts"].items(), key=lambda x: -x[1]):
        print(f"    {k or '(blank)':35s}  {n}")
    print()
    print("  Per-driver ATL:")
    for d, t in sorted(data["by_driver"].items(), key=lambda x: -x[1]["revenue"]):
        full = NAME_MAP.get(d, f"<UNMAPPED:{d}>")
        gross = t["revenue"] - t["carrier"]
        print(f"    {full:25s} ({d:8s})  loads={t['count']}  rev=${t['revenue']:>10,.2f}  car=${t['carrier']:>10,.2f}  gp=${gross:>9,.2f}")

    print()
    print("-" * 60)
    print("  Paste this into src/App.jsx (replace existing ATL_BILLING):")
    print("-" * 60)
    print(f"const ATL_BILLING = {{")
    print(f'  asOf: "{data["as_of"]}",')
    print(f'  loads: {data["loads"]},')
    print(f'  revenue: {data["revenue"]:.2f},      // sum of Invoice Amount')
    print(f'  carrierPay: {data["carrier_pay"]:.2f},   // sum of Carrier Amount (COGS)')
    print(f'  grossProfit: {data["gross_profit"]:.2f},')
    print(f'  grossMargin: {data["gross_margin"]},      // %')
    print(f'  byDriver: [')
    for d, t in sorted(data["by_driver"].items(), key=lambda x: -x[1]["revenue"]):
        full = NAME_MAP.get(d, f"<UNMAPPED:{d}>")
        gross = t["revenue"] - t["carrier"]
        print(f'    {{ name: "{full}",{" "*(28-len(full))}short: "{d}",{" "*(10-len(d))}loads: {t["count"]}, revenue: {t["revenue"]:.2f}, carrier: {t["carrier"]:.2f}, gross: {gross:.2f} }},')
    print(f"  ],")
    print(f"}};")


if __name__ == "__main__":
    f = find_billing_file()
    print(f"Parsing: {os.path.basename(f)}")
    data = parse(f)
    emit(data)
