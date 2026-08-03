"""
Parse the Samsara Vehicle Mileage XLSX dropped into incoming-freightiq/ and
emit the MILES + TRUCK_COUNT + TRUCK_MILES[] constants to paste into src/App.jsx.

Usage:
  python scripts/parse_samsara_mileage.py

Expected file: "Vehicle Mileage - <date range>.xlsx" with columns:
    Vehicle | Jurisdiction | Distance (mi) | Toll Distance (mi)

Local vs Regional split: NV = local, everything else = regional.
"""
import os, sys, glob, csv, re
import openpyxl
from collections import defaultdict

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INCOMING = os.path.join(BASE, "incoming-freightiq")


def find_file():
    # Samsara switched this export from .xlsx to .csv on the Jul 30 2026 drop.
    # Same columns either way (Vehicle / Jurisdiction / Distance / Toll Distance),
    # so accept both and let the newest file win.
    stems = ["Vehicle Mileage*", "*Vehicle Mileage*", "*Samsara*"]
    matches = []
    for stem in stems:
        for ext in (".xlsx", ".csv"):
            matches += glob.glob(os.path.join(INCOMING, stem + ext))
    matches = list(set(matches))
    if not matches:
        print(f"No Samsara mileage export in {INCOMING}/ (looked for 'Vehicle Mileage*.xlsx' / '*.csv')")
        sys.exit(1)
    return max(matches, key=os.path.getmtime)


def _read_rows(path):
    """Yield data rows (header skipped) from either the .xlsx or .csv export."""
    if path.lower().endswith(".csv"):
        # utf-8-sig: the Samsara CSV ships with a BOM on the Vehicle header.
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for i, row in enumerate(csv.reader(fh)):
                if i == 0:
                    continue
                yield tuple(row)
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield row


def parse(path):
    by_truck_state = defaultdict(lambda: defaultdict(float))
    rows = _read_rows(path)
    for r in rows:
        if not r or len(r) < 3:
            continue
        veh, juris, dist = r[0], r[1], r[2]
        if not veh:
            continue
        veh_s = str(veh).strip()
        juris_s = str(juris).strip().upper() if juris else ""
        try:
            d = float(dist) if dist not in (None, "") else 0.0
        except (TypeError, ValueError):
            d = 0.0
        if not juris_s:
            continue
        by_truck_state[veh_s][juris_s] += d

    trucks = []
    fleet_total = 0.0
    fleet_local = 0.0    # NV
    fleet_regional = 0.0 # everything else
    for veh, states in by_truck_state.items():
        local = states.get("NV", 0.0)
        total = sum(states.values())
        regional = total - local
        trucks.append({
            "truck": veh,
            "local": round(local, 1),
            "regional": round(regional, 1),
            "miles": round(total, 1),
            "states": {st: round(d, 1) for st, d in sorted(states.items(), key=lambda x: -x[1])},
        })
        fleet_total += total
        fleet_local += local
        fleet_regional += regional

    trucks.sort(key=lambda t: -t["miles"])
    return {
        "trucks": trucks,
        "fleet_total": round(fleet_total, 1),
        "fleet_local": round(fleet_local, 1),
        "fleet_regional": round(fleet_regional, 1),
        "truck_count": len(trucks),
    }


def emit(data):
    print("-" * 70)
    print(f"  SAMSARA MILEAGE — {data['truck_count']} trucks")
    print("-" * 70)
    print(f"  Fleet total:    {data['fleet_total']:>14,.1f} mi")
    print(f"  Local (NV):     {data['fleet_local']:>14,.1f} mi  ({data['fleet_local']/data['fleet_total']*100:.1f}%)")
    print(f"  Regional:       {data['fleet_regional']:>14,.1f} mi  ({data['fleet_regional']/data['fleet_total']*100:.1f}%)")
    print()
    print(f"  Top 5 trucks:")
    for t in data["trucks"][:5]:
        print(f"    Truck {t['truck']:>4s}  {t['miles']:>10,.1f} mi  ({len(t['states'])} states)")
    print()
    print("-" * 70)
    print("  Paste this into src/App.jsx (replace existing block):")
    print("-" * 70)
    print(f"let MILES     = {data['fleet_total']};     // Samsara Vehicle Mileage report — {data['truck_count']} trucks")
    print(f"let TRUCK_COUNT = {data['truck_count']};       // Active fleet trucks (from Vehicle Mileage report)")
    print(f"let FLEET_LOCAL    = {data['fleet_local']};   // NV miles")
    print(f"let FLEET_REGIONAL = {data['fleet_regional']};   // non-NV miles")
    print()
    print("let TRUCK_MILES = [")
    for t in data["trucks"]:
        states_str = "{" + ",".join(f'"{st}":{d}' for st, d in t["states"].items()) + "}"
        print(f'  {{ truck:"{t["truck"]}", local:{t["local"]}, regional:{t["regional"]}, miles:{t["miles"]}, states:{states_str} }},')
    print("];")


MONTH_DIR = os.path.join(INCOMING, "_monthly-samsara")
_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def emit_monthly():
    """Emit MONTHLY_MILES[] rows from per-month exports in _monthly-samsara/.

    Kept in a SUBFOLDER on purpose: find_file() picks the newest mtime match,
    and the monthly files share a timestamp with the YTD export — left side by
    side, a one-month file can win and silently write April-only data into MILES.

    These rows include EVERY truck (ATL included), matching TRUCK_MILES[]. Only
    the CPM constant MILES is ATL-carved.
    """
    if not os.path.isdir(MONTH_DIR):
        return
    found = []
    for name in os.listdir(MONTH_DIR):
        if not name.lower().endswith((".csv", ".xlsx")):
            continue
        m = re.search(r"Vehicle Mileage - (%s)\b" % "|".join(_MONTHS), name)
        if m:
            found.append((_MONTHS.index(m.group(1)), m.group(1),
                          os.path.join(MONTH_DIR, name)))
    if not found:
        return
    found.sort()
    print()
    print("-" * 70)
    print("  MONTHLY_MILES rows — splice into the existing array in month order")
    print("-" * 70)
    for _, label, path in found:
        d = parse(path)
        trucks = ",".join(
            '"%s":{l:%s,r:%s,t:%s}' % (t["truck"], t["local"], t["regional"], t["miles"])
            for t in d["trucks"])
        print('  { m:"%s", local:%s, regional:%s, total:%s,' %
              (label, d["fleet_local"], d["fleet_regional"], d["fleet_total"]))
        print('    trucks:{%s} },' % trucks)


if __name__ == "__main__":
    f = find_file()
    print(f"Parsing: {os.path.basename(f)}")
    data = parse(f)
    emit(data)
    emit_monthly()
